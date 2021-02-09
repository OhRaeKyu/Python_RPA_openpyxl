from openpyxl import Workbook

wb = Workbook() # 새 워크북 생성
# ws = wb.active # 현재 활성화된 sheet를 가져옴
ws = wb.create_sheet()  #새로운 Sheet 생성
ws.title = "MySheet"  # sheet의 이름 변경
ws.sheet_properties.tabColor = "ff66ff"

ws1 = wb.create_sheet("YourSheet")  # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번 째 sheet 생성

new_ws = wb["NewSheet"] # dictionary 형태로 접근

print(wb.sheetnames)  #모든 sheet 이름 확인

new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
wb.close()