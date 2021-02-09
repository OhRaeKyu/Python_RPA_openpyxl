from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

# 열의 너비 설정
ws.column_dimensions["A"].width = 5
# 행의 높이 설정
ws.row_dimensions[1].height = 50

# 폰트 스타일 설정
a1.font = Font(color = "FF0000", italic = True, bold = True)
b1.font = Font(color = "CC33FF", name = "Arial", strike = True)
c1.font = Font(color = "0000FF", size = 20, underline = "single")

# 테두리 적용
thin_border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
  for cell in row:
    cell.alignment = Alignment(horizontal = "center", vertical = "center")  # 글자의 정렬 위치 설정
    if cell.column == 1:
      continue
    if isinstance(cell.value, int) and cell.value > 90:
      cell.fill = PatternFill(fgColor = "00FF00", fill_type = "solid")  # 배경 색 설정
      cell.font = Font(color = "FFFFFF")

# 틀 고정
ws.freeze_panes = "B2"
wb.save("sample_style.xlsx")