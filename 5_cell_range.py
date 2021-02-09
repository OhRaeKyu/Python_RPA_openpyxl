from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string
from random import *

wb = Workbook()
ws = wb.active

# 1줄 씩 데이터 추가
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
  ws.append([i, randint(0, 100), randint(0, 100)])

# col_B = ws["B"] # B column 데이터 ("영어")
# for cell in col_B:
#   print(cell.value)

# col_range = ws["B:C"] # B 부터 C column 데이터 가져오기
# for cols in col_range:
#   for cell in cols:
#     print(cell.value)

# row_title = ws[1] # 1번 째 row 만 가지고 오기
# for cell in row_title:
#   print(cell.value)

# row_range = ws[2:6] # 2번 째 부터 6번 째 줄까지 가지고 오기
# for rows in row_range:
#   for cell in rows:
#     print(cell.value, end = " ")
#   print()

# row_range = ws[2:ws.max_row]
# for rows in row_range:
#   for cell in rows:
#     # print(cell.value, end = " ")
#     # print(cell.coordinate, end = " ")
#     xy = coordinate_from_string(cell.coordinate)  # cell의 위치를 튜플 형태로 알파벳과 숫자를 나누어 표현
#     print(xy, end = " ")
#   print()

# #######--------- 튜플로 다루기 ---------#######
# print(tuple(ws.rows)) # 전체 rows

# print(tuple(ws.columns)) # 전체 colums
# for row in tuple(ws.rows):  # 전체 row에서
#   print(row[0].value) # 행의 위치가 첫번 째인 값을 출력
# for row in ws.iter_rows(min_row = 1, max_row = 5):  # 최소 row 부터 최대 row 까지
#   print(row[0].value)


# for column in tuple(ws.columns):  # 전체 col에서
#   print(column[0].value)  # 열의 위치가 첫번 째인 값을 출력
# for column in ws.iter_cols(min_col = 1, min_col = 3): # 최소 col 부터 최대 col 까지
#   print(column[0].value)

for row in ws.iter_rows(min_row = 2, max_row = 11, min_col = 2, max_col = 3):
  print(row)

for col in ws.iter_cols(min_row = 1, max_row = 5, min_col = 1, max_col = 3):
  print(col)

wb.save("sample.xlsx")  