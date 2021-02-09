from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

ws.move_range("B1:C11", rows = 0, cols = 1) #'B1부터 C11까지 범위 전체를 cols = 1 만큼 이동
ws["B1"].value = "국어"


wb.save("sample_move.xlsx")