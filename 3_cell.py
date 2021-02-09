from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = "ORK Sheet"

ws["A1"] = 1  # A1셀에 1 이라는 값을 입력
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"].value) # A1 셀의 값을 출력
print(ws.cell(column = 1,row = 1).value)  # column 1 row 1 == ws["A1"]
print(ws.cell(column = 2, row = 1).value) # column = 2 row 1 == ws["B2"]

c = ws.cell(column = 3, row = 1, value = 10) # == ws["C1"] = 10 와 동일
print(c.value)

# 반복문을 이용한 랜덤 숫자 채우기
for x in range(1, 11):
  for y in range(1, 11):
    ws.cell(column = x, row = y, value = randint(0, 100))

wb.save("sample.xlsx")