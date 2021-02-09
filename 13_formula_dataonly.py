from openpyxl import load_workbook

# # 수식 그대로 가져오고 있음
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# for row in ws.values:
#   for cell in row:
#     print(cell)

# 수식이 계산 된 데이터를 불러오기
wb = load_workbook("sample_formula.xlsx", data_only = True)
ws = wb.active

for row in ws.values:
  for cell in row:
    print(cell)